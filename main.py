import streamlit as st
from tqdm import tqdm
import requests
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
import time
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import re


class PiguAPIError(Exception):
    pass


class PiguClient:
    def __init__(self, username: str, password: str, seller_id: str):
        self.username = username
        self.password = password
        self.seller_id = seller_id
        self.token = None

    def _login(self) -> str:
        headers = {"Content-Type": "application/json"}
        body = {"username": self.username, "password": self.password}

        try:
            response = requests.post(
                "https://pmpapi.pigugroup.eu/v3/login",
                headers=headers,
                json=body,
                timeout=30
            )
            response.raise_for_status()

            token = response.json().get("token")
            if not token:
                raise PiguAPIError("Token d'authentification manquant")
            return token

        except requests.RequestException as e:
            raise PiguAPIError(f"Erreur d'authentification : {e}")

    def fetch_offers(self, page_size: int = 100) -> List[Dict]:
        if not self.token:
            self.token = self._login()

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Pigu-mp {self.token}"
        }

        params = {
            "amount_from": "1",
            "page_size": page_size
        }

        url = f"https://pmpapi.pigugroup.eu/v3/sellers/{self.seller_id}/offers"
        offers_data = []

        st.write("\n🔄 Récupération des offres en cours...")
        progress_bar = st.progress(0)

        page_count = 0
        max_pages = 100  # Valeur arbitraire pour normaliser, ou à adapter si vous avez un max connu

        with tqdm(total=float('inf'), desc='Progression', unit=' offres') as pbar:
            while True:
                try:
                    response = requests.get(url, headers=headers, params=params, timeout=30)
                    response.raise_for_status()

                    data = response.json()
                    offers_batch = data.get("offers", [])
                    offers_data.extend(offers_batch)

                    page_count += 1
                    progress = min(page_count / max_pages, 1.0)
                    progress_bar.progress(progress)

                    pbar.update(len(offers_batch))
                    pbar.set_postfix({'Total': len(offers_data)})

                    next_url = data.get("meta", {}).get("next")
                    if not next_url:
                        break

                    url = next_url
                    time.sleep(1)  # Limiter les requêtes

                except requests.RequestException as e:
                    raise PiguAPIError(f"Erreur de récupération des offres : {e}")

        st.write(f"\n✅ Récupération terminée ! {len(offers_data)} offres trouvées.")
        progress_bar.progress(1.0)
        return offers_data

    def save_offers_to_excel(self, filename: str):
        st.write("\n📊 Traitement des données...")

        offers = self.fetch_offers()
        df = pd.json_normalize(offers)

        st.write("📋 Aperçu des offres récupérées :")
        st.dataframe(df)

        try:
            df.to_excel(
                filename,
                index=False,
                freeze_panes=(1, 0),
                engine='openpyxl'
            )

            st.write("💼 Mise en forme du tableau...")
            wb = load_workbook(filename=filename)
            ws = wb.active

            header_style = PatternFill(
                start_color='001944',
                end_color='001944',
                fill_type='solid'
            )

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_style
                cell.font = Font(color='FFFFFF')

                ws.row_dimensions[1].height = 25
                column_letter = get_column_letter(col_idx)
                column_width = max(len(str(cell.value)) + 2, 15)
                ws.column_dimensions[column_letter].width = column_width

            wb.save(filename)
            st.success("✨ Sauvegarde terminée avec succès !")

        except Exception as e:
            st.error(f"Erreur lors de la sauvegarde Excel : {str(e)}")
            raise PiguAPIError(f"Erreur lors de la sauvegarde Excel : {str(e)}")


# Configuration de l'interface Streamlit
st.title("Extraction des Offres PHH")
st.sidebar.header("Paramètres de Connexion")

# Champs de saisie pour les informations d'identification
seller_id = st.sidebar.text_input("Identifiant Vendeur", value="")
username = st.sidebar.text_input("Nom d'utilisateur", value="")
password = st.sidebar.text_input("Mot de passe", type="password")

# Bouton de déclenchement
if st.sidebar.button("Extraire les Offres"):
    try:
        client = PiguClient(username, password, seller_id)
        now = datetime.now()
        date_time = now.strftime('%Y-%m-%d_%H-%M')

        match = re.search(r'\d+', username)
        bzp_seller_id = match.group()

        filename = f"PHH_Offres_{bzp_seller_id}_{date_time}.xlsx"
        client.save_offers_to_excel(filename)

    except PiguAPIError as e:
        st.error(f"Erreur : {str(e)}")
